from openpyxl import load_workbook
from db import get_connection

def _clean(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()

def _num(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0

def import_payroll_excel(file_path: str) -> str:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    title = _clean(ws.cell(2, 1).value) or _clean(ws.title)

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM payroll_rows")
    cur.execute("DELETE FROM payroll_imports")
    conn.commit()

    rows_count = 0
    total_gross = 0.0
    total_deductions = 0.0
    total_net = 0.0
    total_salary_advance = 0.0
    total_bank_loan = 0.0

    imported_at = __import__("datetime").datetime.now().isoformat(timespec="seconds")
    cur.execute(
        "INSERT INTO payroll_imports(source_file, imported_at, payroll_month) VALUES (?, ?, ?)",
        (file_path, imported_at, title)
    )
    import_id = cur.lastrowid

    for row in range(6, ws.max_row + 1):
        name = _clean(ws.cell(row, 2).value)
        if not name:
            continue
        normalized_name = name.replace("ـ", "").strip()
        if "الاجمال" in normalized_name or "الإجمال" in normalized_name:
            break

        employee_no = ws.cell(row, 1).value
        if employee_no in (None, ""):
            continue
        gross = _num(ws.cell(row, 18).value)
        deductions = _num(ws.cell(row, 30).value)
        net = _num(ws.cell(row, 31).value)
        salary_advance = _num(ws.cell(row, 24).value)
        bank_loan = _num(ws.cell(row, 25).value)
        insurance = _num(ws.cell(row, 22).value)
        tax = _num(ws.cell(row, 29).value)

        cur.execute("""
            INSERT INTO payroll_rows(
                import_id, employee_no, employee_name, payroll_month,
                gross_total, total_deductions, net_pay,
                salary_advance_installment, bank_loan_installment,
                insurance_employee, tax_amount
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            import_id, int(employee_no) if employee_no not in (None, "") else None, name, title,
            gross, deductions, net, salary_advance, bank_loan, insurance, tax
        ))

        rows_count += 1
        total_gross += gross
        total_deductions += deductions
        total_net += net
        total_salary_advance += salary_advance
        total_bank_loan += bank_loan

    cur.execute("""
        UPDATE payroll_imports
        SET employees_count=?, gross_total=?, deductions_total=?, net_total=?,
            salary_advance_total=?, bank_loan_total=?
        WHERE id=?
    """, (rows_count, total_gross, total_deductions, total_net, total_salary_advance, total_bank_loan, import_id))
    conn.commit()
    conn.close()

    return (
        f"تم استيراد ملف المرتبات بنجاح\n"
        f"الفترة: {title}\n"
        f"عدد العاملين: {rows_count}\n"
        f"إجمالي الجملة: {total_gross:,.2f}\n"
        f"إجمالي المستقطعات: {total_deductions:,.2f}\n"
        f"صافي المستحق: {total_net:,.2f}\n"
        f"إجمالي قسط السلف: {total_salary_advance:,.2f}\n"
        f"إجمالي قسط سلفة البنك: {total_bank_loan:,.2f}"
    )
